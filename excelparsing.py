# -*- coding: utf-8 -*-
"""
Created on Tue Feb 14 09:57:33 2017

@author: LELJD420
"""

import openpyxl
from os.path import join, dirname, abspath
import numpy as np



class xlimpexp:
    
    def __init__(self, importedfilename='file.xlsx', exportedfilename='output.xlsx'):
        self.Ifn, self.Efn = importedfilename, exportedfilename
    
    def _gensheet(self):
        book = openpyxl.Workbook()
        sh = book.active
        self.SHn = sh
        return sh, book
        
    def exportdictlist(self, d):
        """
        export dictionary of lists as headed columns in excel
        """
        sh, book = self._gensheet()
        for col, k in enumerate(d):
            sh.write(0, col, list(d.keys())[col])
            for row, item in enumerate(d[k]):
                sh.write(row+1, col, item)
        book.save(self.Efn)
        
    def exportdictlistarray(self, d):
        """
        export dictionary of lists of arrays of equal length as headed columns in excel
        """
        sh, book = self._gensheet()
        n = len(list(d.keys()))
        for col, k in enumerate(d):
            for row, item in enumerate(d[k]):
                sh.write(0, n*col, list(d.keys())[col])
                for i in np.arange(0,len(item)):
                    sh.write(row+1, n*col+i, item[i])
        book.save(self.Efn)
    
    def exportdictdict(self, dd):
        """
        export dictionary of dictionaries as headed columns in excel
        """
        sh, book = self._gensheet()
        for head, kdd in enumerate(dd):
            offset = head*len(dd[kdd])
            sh.write(0, offset, kdd)
            for col, kd in enumerate(dd[kdd]):
                sh.write(1, offset+col, kd)
                for row, item in enumerate(dd[kdd][kd]):
                    sh.write(row+2, col, item)
        book.save(self.Efn)
     
    def exportlist(self, l, startcell='A1', header='data'):
        sheet, book = self._gensheet()
        sheet[startcell] = header
        col, row = openpyxl.utils.coordinate_from_string(startcell)
        for i, el in enumerate(l):
            sheet[col+str(row+i+1)] = el        
        book.save(self.Efn)
        
        
    def idx_to_str(self, idx_coords):
        col, row = idx_coords[0], idx_coords[1]
        str_coords = openpyxl.utils.get_column_letter(col) + str(row)
        return str_coords
                
    def importcolumn(self, stringnames, sheet_index=0, data_ordering='columns'):
        """
        Imports columns or rows of given header and returns a dictionary of coords.
        stringnames is a list of strings
        """
        try:
            filepath = join(dirname(abspath(__file__)), self.Ifn)
            excelfile = openpyxl.load_workbook(filepath, data_only=True)
            sheetnames = excelfile.sheetnames
            sheet = excelfile[sheetnames[sheet_index]]
            tInitPositionDict, tEndPositionDict = {}, {}
            
            if data_ordering == 'columns':
                data_ord_idx = 1
            else:
                data_ord_idx = 0

            meth_list = ['column', 'row']
            # find headers' coords
            cell_iterator = tuple(getattr(sheet, data_ordering))
#            print(data_ordering)
            
            for s in stringnames:
                tInitPositionDict[s], tEndPositionDict[s] = [], []
                for arr in cell_iterator:
#                    print(arr)
                    for cell in arr:
                        if cell.value == s:
                            
                            idx_coords = [cell.col_idx, cell.row]
#                            print(idx_coords)
                            idx_coords[data_ord_idx] += 1
#                            print(idx_coords)
                            list_arr = list(arr)
#                            print(list_arr[-1])
                            tInitPositionDict[s].append(list_arr[idx_coords[data_ord_idx]-1])
                            tEndPositionDict[s].append(list_arr[-1])
#                            print(list_arr[idx_coords[data_ord_idx]:])
                            for c in list_arr[idx_coords[data_ord_idx]:]:
#                                print(list_arr[c].value)
                                if type(c.value) == str:
#                                    print('string',c.value)
                                    index = openpyxl.utils.column_index_from_string(getattr(c, meth_list[data_ord_idx])) - 2
#                                    print(index)
                                    tEndPositionDict[s][-1] = list_arr[index]
                                    break
                        
                            
                            
            print(tInitPositionDict,tEndPositionDict)  

            # actual data
            data = {}
            for s in tInitPositionDict:
                data[s] = []
                for pos_cell_init_idx in range(len(tInitPositionDict[s])):
                    cell_range = tInitPositionDict[s][pos_cell_init_idx].coordinate + ':' + tEndPositionDict[s][pos_cell_init_idx].coordinate
#                    print(cell_range)
                    slice_data = []
#                    print(sheet[cell_range])
                    if data_ordering == 'columns':
                        for c in sheet[cell_range]:
                            slice_data.append(c[0].value)
                    else:
                        for c in sheet[cell_range][0]:
                            slice_data.append(c.value)
                        
                    data[s].append(slice_data)
                for j in range(len(data[s])):
                    data[s][j] = [x for x in data[s][j] if x is not None]
            return data
            
        except FileNotFoundError:
            print('No .xls found, use provided data array instead')
            raise
        except IndexError:
            print('No data found')
            raise
            
    def importtxt (self, separator=';'):
        """
        Import text files as matrix
        """
        try:
            filepath = join(dirname(abspath(__file__)), self.Ifn)
            file = open(filepath, "r")
            lines = file.readlines()
            split_lines = []
            for l in lines:
                if l[-1] == '\n':
                    l = l[:-1]
                split_lines.append(l.split(separator))
                
                for i in range(len(split_lines[-1])):
                    try:
                        split_lines[-1][i] = float(split_lines[-1][i])
                    except ValueError:
                        continue
            
        except FileNotFoundError:
            print('File not found, try again.')  
            raise
        
        #sh, book = self._gensheet()
        
        return split_lines
            